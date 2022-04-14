'''Задачи 1-6'''
from openpyxl import *

wb = load_workbook('food.xlsx')
sheet = wb.worksheets[0]
firstColumn = sheet['A']
theValue = sheet['F19'].value
myRange = sheet['A18:H18']
print(myRange)
myListExcel = []
for item in myRange[0]:
    myListExcel.append(item.value)
print(myListExcel)
for value in myListExcel:
    print(value)

'''Задача 7
Сделайте перебор данных установив лимит в 5 рядов из myListExcel'''

print('\nзадача 7 ****************** ')
for value in myListExcel[0:5]:
    print(value)

'''Задача 8
Попросить вести пользователя следующие данные, как в след. 
примере:
[‘Имя’,’Возраст’,’Зарплата’,’Год работ’],
[‘Самат’, ‘Улан’]
[25, 30]
[25000, 32000]
[5,7]'''

tableHeader = ['Имя', 'Возраст', 'Зарплата', 'Год работ']
wb = Workbook()
sheet = wb.worksheets[0]
counter = 1
for title in tableHeader:
    sheet.cell(row=1, column=counter).value = title
    counter += 1
sheet.cell(row=2, column=1).value = input('Введите имя: ')
sheet.cell(row=2, column=2).value = input('Введите взраст: ')
sheet.cell(row=2, column=3).value = input('Введите ЗП: ')
sheet.cell(row=2, column=4).value = input('Введите стаж: ')
wb.save('workers.xlsx')
