#import openpyxl
from openpyxl import *
from openpyxl.styles import *

# ЧТЕНИЕ ДАННЫХ ИЗ ЭКЗЕЛЬ
# загрузка экзель и 1-ой страницы
wb = load_workbook('example12.xlsx')
sheet = wb.worksheets[0] #1,2,3

#чтение по ячейке
namePerson1 = sheet['A4'].value
agePerson1  = sheet['B4'].value
agePerson2  = sheet['B3'].value

print(namePerson1)
print(agePerson1)

print(type(namePerson1))
print(type(agePerson1))

resultage = agePerson1 + agePerson2
print(resultage)

cells = sheet['A6':'C6']
print(cells)

for el1, el2, el3 in cells:
    print(el1.value,el2.value,el3.value)

for el in cells[0]:
    print(el.value, end=" ")

namePerson2 = sheet['B4']
print(namePerson2)

# tupleSome = ((1,3,4),(23,4,5))
# print(tupleSome[0][1])
# print(cells[0][1])

zpSome = sheet.cell(row=4, column=3).value
print(zpSome)
zpSome2 = sheet.cell(row=5, column=3).value
print(zpSome2)


salaryList = sheet['C']
salaryList = salaryList[1:]
sumSalary = 0
for salary in salaryList:
    print(salary.value, end=" ")
    sumSalary += salary.value

print(f'Sum salary is {sumSalary}')


# ЗАПИСЬ ДАННЫХ В ЭКЗЕЛЬ
countryList = ['Kyrgyzstan', 'Russia', 'USA', 'Uzbekistan', 'China', 'France']
capitalList = ['Bishkek', 'Moscow', 'Washington', 'Tashkent', 'Beijing', 'Paris']
vvp = [1000, 10000, 15000, 2000, 25000, 12000]

#Создание нового файла
wb2 = Workbook()
sheet2 = wb2.worksheets[0]

cell1Header = sheet2.cell(row=1, column=1).value = 'Страна'
cell2Header = sheet2.cell(row=1, column=2).value = 'Столица'
cell3Header = sheet2.cell(row=1, column=3).value = 'ВВП'

sheet2['A1'].font = Font(bold=True)
sheet2['B1'].font = Font(bold=False)
sheet2['C1'].font = Font(bold=True)

# sheet2.cell(row=2, column=1).value = countryList[0]
# sheet2.cell(row=3, column=1).value = countryList[1]
# sheet2.cell(row=4, column=1).value = countryList[2]
# sheet2.cell(row=5, column=1).value = countryList[3]

counter = 2
for i in range(6):
    sheet2.cell(row=counter, column=1).value = countryList[i]
    sheet2.cell(row=counter, column=2).value = capitalList[i]
    sheet2.cell(row=counter, column=3).value = vvp[i]
    counter += 1

wb2.save('countryInfo.xlsx')
print('We saved!')



