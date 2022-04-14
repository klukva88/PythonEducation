from openpyxl import *
# Чтение название всех страниц Excel

wb = load_workbook('food.xlsx')
#sheetNames = wb.get_sheet_names()
sheetNames = wb.sheetnames
#sheet = wb[sheetNames[4]]
sheet = wb['delivery_maninfo']

print(sheetNames)
print(sheet)

cell = sheet['A3':'D3']
for i in cell[0]:
    print(i.value, end=" ")
print('\n')

#Первый способ чтение большого промежутка
sheet2 = wb['Food_List']
cell2 = sheet2['C20':'G45']

"""
listSome = [
    [234,54,65,6],
    [234,54,65,6],
    [234,54,65,6]
]

for row in listSome:
    for num in row:
        print(num)
"""
print('*'*20)
# for row in cell2:
#     for data in row:
#         print(data.value, end=" ")
#     print(end="\n")

#Второй способ чтение большого промежутка
#(min_row - start row, min_col - start col, max_row - end row, max_col - end col)
cell3 = sheet2.iter_rows(min_row=49, min_col=2, max_row=60, max_col=5)
print('*'*20)

for row in cell3:
    for data in row:
        print(data.value, end=" ")
    print(end="\n")

print('*'*20)

rows = sheet2.rows
dataList = []
print('*'*20)
for i in rows:
    for data in i:
        print(data.value, end=" ")
    print(end="\n")

cols = sheet2.columns
dataList = []
print('*'*20)
for i in cols:
    for data in i:
        print(data.value, end=" ")
    print(end="\n")


sheet3 = wb['delivery_maninfo']
sheet3.merge_cells('B6:C6')
cell4 = sheet3['B6'].value = 'someValue'
wb.save('food.xlsx')



sheet4 = wb['numbers']
sheet4['A1'] = 65
 # Установить высоту строки
sheet4.row_dimensions[1].height = 50
 # Установить высоту столбца
sheet4.column_dimensions['A'].width = 30
wb.save('food.xlsx')

# sheet5 = wb['delivery_adress']
# wb.remove(sheet5)

sheet2 = wb['Food_List']
print(sheet2['C16'].value)
sheet2['C16'] = 'Bishkek'

wb.save('food.xlsx')


#Сохранение одним потоком
wb2 = Workbook()
sheet6 = wb2.create_sheet('Student_info')

sheet7 = wb2['Sheet']
wb2.remove(sheet7)
djangoCourseInfo =(
    ('Name', 'Country', 'Current_work'),
    ('Aselya', 'Russia', 'Buhgalter'),
    ('Dinara', 'Poland', 'Math'),
    ('Tilek', 'Kyrgyzstan', 'Bank Worker'),
    ('Nikolai', 'Russia', 'Engineer'),
    ('Aizhan', 'Kyrgyzstan', 'Buhgalter'),
    ('Aiperi', 'Usa', 'Student'),
)

for data in djangoCourseInfo:
    sheet6.append(data)

wb2.save('Student.xlsx')



