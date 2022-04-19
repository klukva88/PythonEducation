from openpyxl import *

'''Задача 1
Чтение данных в промежутке ['B12':'G28'] 2 способами :
a) По координатам
b) По указаниям строки и колонок'''
wb = load_workbook('food.xlsx')
sheet = wb.worksheets[0]
cells = sheet['B12':'G28'] #Чтение по координатам
cells2 = sheet.iter_rows(min_row=12, min_col=2, max_row=28, max_col=7) #Чтение по указателям строк в колонке

for cell in cells:
    for value in cell:
        print(value.value, end=' ')
print('\n')
for cell in cells2:
    for value in cell:
        print(value.value, end=' ')

'''Задача 2
Запись данных в файл экзель одним потоком, т.е. путем кортежа
информацию о некой семье.'''

family = (
    ('Имя', 'Статус', 'Возраст', 'Вес', 'Доход'),
    ('Степан Васильев', 'Отец', 55, 85, 50000),
    ('Ольга Иванова', 'Мать', 53, 72, 32000),
    ('Григорий Васильев', 'Сын', 22, 80, 35000),
    ('Наталья Васильева', 'Дочь', 19, 65, 20000),
    ('Петр Васильев', 'Сын', 18, 75, 10000)
)

familyWb = Workbook()
sheetToRemove = familyWb.worksheets[0]
sheetTask2 = familyWb.create_sheet('Family')
for member in family:
    sheetTask2.append(member)
familyWb.remove(sheetToRemove)
familyWb.save('family_members.xlsx')

'''Задача 3
Необходимо добавить дополнительную строку «ИТОГО» и 
посчитать итоговые значения для колонок с помощью формул.'''

wb = load_workbook('family_members.xlsx')
sheet = wb.active
total = (
    ('ИТОГО', None, '=SUM(C2:C6)', '=SUM(D2:D6)', '=SUM(E2:E6)')
)
sheet.append(total)
wb.save('family_members.xlsx')


