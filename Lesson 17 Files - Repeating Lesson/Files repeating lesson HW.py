'''Задача 1
Скопируйте содержание одного файла в другой файл'''
def copyFile(file1, file2):
    with open(file1, 'r') as fileOne:
       fileOne = fileOne.read()
    with open(file2, 'w') as fileTwo:
        fileTwo.write(fileOne)
#copyFile('text.txt', 'text_copy.txt')

'''Задача2
Отобразите последние 3 строки из файла'''

def lastTreeLines(file):
    with open(file, 'r', encoding='utf8') as file:
        lines = file.readlines()
        for line in lines[-3::]:
            print(line.rstrip())

#lastTreeLines('text_copy.txt')

def minWordLength(file):
    with open(file, 'r', encoding='utf8') as file:
        lines = file.read().replace('\n', ' ').split(' ')
        print(lines)
        minLength = sorted(lines, key=lambda line: len(line))
        print(f'Слово наименьшей длины - {minLength[0]}')
#minWordLength('text_copy.txt')

'''Задача 4
Из файла “food” выведите только 3 колонку'''
from openpyxl import *
def thirdColumn(file):
    wb = load_workbook(file)
    sheet = wb.worksheets[0]
    column = sheet['C']
    for item in column:
        print(item.value)

#thirdColumn('food.xlsx')

'''Задача 5
Из файла “food” выведите только 5 строку'''

def fifthRow(file):
    wb = load_workbook(file)
    sheet = wb.worksheets[0]
    row = sheet.iter_rows(min_row=5)
    print(row)
    for item in row:
        for cell in item:
            print(cell.value)
#fifthRow('food.xlsx')

'''Задача 6
Примите 2 координаты для отображения промежутка от 
пользователя и отобразите этот промежуток'''


def showRange(cordinate_1, coordinate_2):
    wb = load_workbook('food.xlsx')
    sheet = wb.worksheets[0]
    cellRange = sheet[cordinate_1 : coordinate_2]
    for item in cellRange:
        for cell in item:
           print(cell.value)
#showRange(input('Введите координату 1: '), input('Введите координату 2: ') )

'''Задача 7
Примите от пользователя название города и отобразите данные 
только этого города из файла “food”'''

def showCityData(city):
    wb = load_workbook('food.xlsx')
    sheet = wb.worksheets[0]
    selectedRows = []
    for row in sheet.rows:
        for cell in row:
            if cell.value == city:
                selectedRows.append(row)
    cityData = [[item.value for item in data] for data in selectedRows]
    for rowData in cityData:
        print(rowData, end='\n')

#showCityData(input('Введите город: '))

'''Задача 8
Создайте новую страницу в файле food.xlsx и скопируйте 
промежуток «B14:G27'''
wb = load_workbook('food.xlsx')

def createNewPage(listName):
    wb.create_sheet(listName, 0)
    sheets = wb.sheetnames
    print(f'Лист {listName} создан!\n'
          f'Список листов: {sheets}')
def copyCellRange(startCell, endCell, listSource, endList):
    sheetSource = wb[listSource]
    endSheet = wb[endList]
    print(f'Выбран лист {list}')
    cellRange = [[data.value for data in cell] for cell in sheetSource[startCell:endCell]]
    print(cellRange)
    for value in cellRange:
        endSheet.append(value)
    wb.save('food.xlsx')

newSheet = input('Введите название нового листа: ')
createNewPage(newSheet)

selectSource = input('Выберете лист откуда скопировать диапазон: ')
startCell = input('Введите первую ячейку диапазона: ')
endCell = input('Введите последнюю ячейку диапазона: ')
selectEndList = input('Выберете лист куда вставить диапазон: ')

copyCellRange(startCell, endCell, selectSource, selectEndList)